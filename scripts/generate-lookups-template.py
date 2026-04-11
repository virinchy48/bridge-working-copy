#!/usr/bin/env python3
"""
Generate the Lookup Values Excel template that ships with the BIS app.

Pulls every active Lookup row from the running CDS server (default
http://localhost:4044) and writes a single-sheet xlsx with:
  - Frozen header row
  - Bold header
  - Column widths sized to content
  - Column comments explaining each field
  - One row per existing lookup, grouped/sorted by category

The output file is the starter template the user downloads from the
Mass Upload → Lookup Values → "Download Template" button. Users can edit
it in Excel, save as CSV/XLSX, and re-upload via the same screen.

Run via:
    python3 scripts/generate-lookups-template.py
or with custom URL:
    LOOKUPS_URL=http://other:4044/bridge-management/Lookups python3 scripts/generate-lookups-template.py
"""

import os
import sys
import json
import urllib.request
import base64
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter

OUT_PATH = os.path.join(
    os.path.dirname(os.path.abspath(__file__)),
    "..", "app", "bridge-management", "webapp", "resources", "templates",
    "lookups-template.xlsx"
)

URL = os.environ.get(
    "LOOKUPS_URL",
    "http://localhost:4044/bridge-management/Lookups?$top=10000&$orderby=category,displayOrder,code"
)
USER = os.environ.get("LOOKUPS_USER", "admin")
PWD  = os.environ.get("LOOKUPS_PWD", "admin")


def fetch_lookups():
    auth = base64.b64encode(f"{USER}:{PWD}".encode()).decode()
    req = urllib.request.Request(URL, headers={
        "Accept": "application/json",
        "Authorization": f"Basic {auth}"
    })
    with urllib.request.urlopen(req, timeout=10) as r:
        return json.load(r).get("value", [])


def main():
    rows = fetch_lookups()
    if not rows:
        print("WARN: no rows returned — generating empty template", file=sys.stderr)

    wb = Workbook()
    ws = wb.active
    ws.title = "Lookups"

    headers = ["category", "code", "description", "displayOrder", "isActive"]
    comments = {
        "category":     "REQUIRED. Upper-case identifier (max 50 chars). Examples: POSTING_STATUS, CONDITION, SCOUR_RISK. Existing categories used by the UI dropdowns are listed below.",
        "code":         "REQUIRED. Upper-case value the dropdown emits when picked (max 200 chars). Two rows with the same (category, code) will UPDATE the existing row instead of creating a duplicate.",
        "description":  "Optional. Human-readable label shown in the dropdown (max 300 chars). If empty, the code is shown.",
        "displayOrder": "Optional integer. Lower numbers appear first. Use 10/20/30 spacing so you can insert new values later without renumbering.",
        "isActive":     "true | false. Inactive entries are hidden from dropdowns but kept in the database. Defaults to true on insert."
    }

    # Header row with comments
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_align = Alignment(horizontal="left", vertical="center")
    border = Border(bottom=Side(border_style="thin", color="FFFFFF"))

    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border
        cell.comment = Comment(comments[h], "BIS")

    ws.freeze_panes = "A2"

    # Data rows
    light_grey = PatternFill("solid", fgColor="F2F2F2")
    last_category = None
    use_grey = False
    for r_idx, row in enumerate(rows, start=2):
        if row["category"] != last_category:
            use_grey = not use_grey
            last_category = row["category"]
        for c_idx, key in enumerate(headers, start=1):
            v = row.get(key)
            if isinstance(v, bool):
                v = str(v).lower()
            ws.cell(row=r_idx, column=c_idx, value=v)
            if use_grey:
                ws.cell(row=r_idx, column=c_idx).fill = light_grey

    # Column widths sized to content
    widths = {"category": 26, "code": 28, "description": 50, "displayOrder": 14, "isActive": 12}
    for col_idx, h in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = widths.get(h, 20)

    # Add a second sheet with category summary
    summary = wb.create_sheet("Categories")
    summary_headers = ["category", "row_count", "purpose"]
    purpose_map = {
        "ACCESS_METHOD":         "Inspection access method (BridgeDetail / InspectionCreate)",
        "ASSET_CLASS":           "Bridge / culvert / footbridge etc. (BridgeForm)",
        "CAPACITY_STATUS":       "Bridge capacity rating",
        "CONDITION":             "Element/structure condition labels (BridgeForm, filters)",
        "DEFECT_CATEGORY":       "Defect category (Defects)",
        "DEFECT_CLASSIFICATION": "Detailed defect codes D1-D13 (InspectionCreate)",
        "DEFECT_EXTENT":         "How widespread a defect is",
        "DEFECT_PRIORITY":       "Defect priority (Defects)",
        "DEFECT_SEVERITY":       "Defect severity (Defects, InspectionCreate)",
        "DEFECT_STATUS":         "Defect status filter (Defects)",
        "DESIGN_LOAD":           "AS 5100 / AASHTO design load (BridgeForm)",
        "ELEMENT_GROUP":         "Bridge element grouping",
        "EXTERNAL_SYSTEM_TYPE":  "BANC, S/4HANA, RMS, VicRoads etc. (BridgeForm, IntegrationHub)",
        "INSPECTION_STANDARD":   "AS 5100.7, NAASRA BMS etc. (InspectionCreate)",
        "INSPECTION_STATUS":     "Inspection status filter (InspectionDashboard)",
        "INSPECTION_TYPE":       "L1 Routine, L2 Principal etc. (InspectionCreate)",
        "INTERVENTION_TYPE":     "Repair, rehab, replace etc.",
        "MAINTENANCE_URGENCY":   "Maintenance urgency band",
        "MATERIAL":              "Concrete, steel, timber etc.",
        "MEASUREMENT_UNIT":      "t, m, kph, kN etc. (Restrictions, BridgeDetail)",
        "NHVR_APPROVAL_CLASS":   "NHVR PBS approval class (BridgeForm)",
        "PERMIT_DECISION":       "Approve / Deny / Conditions (Permits)",
        "PERMIT_STATUS":         "Permit status filter (Permits)",
        "PERMIT_TYPE":           "Permit type filter (Permits)",
        "POSTING_STATUS":        "Bridge posting status (BridgeForm) — server-validated enum",
        "PROGRAMME_STATUS":      "Investment programme status",
        "RATING_METHOD":         "Bridge rating method",
        "RESTRICTION_DIRECTION": "Both / Increasing / Decreasing / NESW (Restrictions)",
        "RESTRICTION_STATUS":    "Restriction status (Restrictions)",
        "RESTRICTION_TYPE":      "Restriction type (Restrictions)",
        "RISK_BAND":             "Risk band Low/Medium/High/Critical (Bridges filter)",
        "ROUTE_CLASS":           "PBS, HML, B-Double etc. (FreightRoutes)",
        "ROUTE_STATUS":          "Active / Suspended / Under Review (FreightRoutes)",
        "SCOUR_RISK":            "Scour risk (BridgeForm) — server-validated enum",
        "STATE":                 "Australian state codes",
        "STRUCTURAL_ADEQUACY":   "Adequate / Marginal / Deficient (BridgeDetail)",
        "STRUCTURAL_RISK":       "Structural risk band",
        "STRUCTURE_TYPE":        "Beam, Arch, Truss etc. (BridgeForm)",
        "VEHICLE_CLASS":         "B-Double, Road Train, PBS levels (BridgeDetail)",
        "WORK_ORDER_PRIORITY":   "Work order priority (WorkOrders, Defects)",
        "WORK_ORDER_STATUS":     "Work order status (WorkOrders)",
    }
    for col_idx, h in enumerate(summary_headers, start=1):
        cell = summary.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill

    counts = {}
    for r in rows:
        counts[r["category"]] = counts.get(r["category"], 0) + 1
    for r_idx, cat in enumerate(sorted(counts.keys()), start=2):
        summary.cell(row=r_idx, column=1, value=cat)
        summary.cell(row=r_idx, column=2, value=counts[cat])
        summary.cell(row=r_idx, column=3, value=purpose_map.get(cat, ""))
    summary.column_dimensions["A"].width = 26
    summary.column_dimensions["B"].width = 12
    summary.column_dimensions["C"].width = 70
    summary.freeze_panes = "A2"

    os.makedirs(os.path.dirname(OUT_PATH), exist_ok=True)
    wb.save(OUT_PATH)
    print(f"WROTE {OUT_PATH}")
    print(f"  rows:       {len(rows)}")
    print(f"  categories: {len(counts)}")


if __name__ == "__main__":
    main()
